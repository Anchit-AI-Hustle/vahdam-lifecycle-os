#!/usr/bin/env python3
"""Build the VAHDAM Lifecycle OS visual dashboard as an .xlsx (multi-tab + native
charts) that uploads/converts cleanly into Google Sheets.

Catalog tabs are populated from REAL data (data/catalog/products_{us,uk,global}.json).
Lifecycle / Campaign / Competitor tabs are brand-styled TEMPLATES wired for the team
to paste or connect live Supabase/Klaviyo data (this builder has no Supabase access).

Usage: python3 scripts/build_sheets_dashboard.py [out.xlsx]
"""
import json, os, sys, statistics
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "scratchpad", "vahdam_dashboard.xlsx")

# ── Brand palette (style guide) ──────────────────────────────────────────────
GREEN = "004A2B"; GOLD = "AB8743"; INK = "171717"; CREAM = "FBF5EA"
WHITE = "FFFFFF"; GREEN_SOFT = "E4EDE8"; GOLD_SOFT = "F3EADc".upper()

F_TITLE   = Font(name="Calibri", size=22, bold=True, color=GREEN)
F_SUB     = Font(name="Calibri", size=11, color="5B6B63")
F_H       = Font(name="Calibri", size=13, bold=True, color=GREEN)
F_TH      = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_KPI_N   = Font(name="Calibri", size=24, bold=True, color=GREEN)
F_KPI_L   = Font(name="Calibri", size=9, bold=True, color=GOLD)
F_CELL    = Font(name="Calibri", size=10, color=INK)
F_NOTE    = Font(name="Calibri", size=9, italic=True, color="8A8A8A")

FILL_HEAD = PatternFill("solid", fgColor=GREEN)
FILL_KPI  = PatternFill("solid", fgColor=CREAM)
FILL_GOLD = PatternFill("solid", fgColor=GOLD_SOFT)
FILL_ALT  = PatternFill("solid", fgColor=GREEN_SOFT)
THIN = Side(style="thin", color="D8D0BE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def load_catalog(region):
    p = os.path.join(ROOT, "data", "catalog", f"products_{region}.json")
    with open(p) as f:
        data = json.load(f)
    if isinstance(data, dict):
        items = data.get("products") or next((v for v in data.values() if isinstance(v, list)), [])
    else:
        items = data
    return items


def price_of(p):
    try:
        return float(str(p.get("price", "")).replace("$", "").replace("£", "").replace(",", "").strip())
    except (ValueError, AttributeError):
        return None


def region_stats(items):
    prices = [v for v in (price_of(p) for p in items) if v and v > 0]
    return {
        "count": len(items),
        "with_price": len(prices),
        "min": round(min(prices), 2) if prices else 0,
        "avg": round(statistics.mean(prices), 2) if prices else 0,
        "median": round(statistics.median(prices), 2) if prices else 0,
        "max": round(max(prices), 2) if prices else 0,
        "types": Counter(p.get("type", "—") for p in items),
        "caffeine": Counter(p.get("caffeine", "—") for p in items),
        "form": Counter(p.get("form", "—") for p in items),
        "prices": prices,
    }


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEAD; cell.font = F_TH; cell.alignment = CENTER; cell.border = BORDER


def banner(ws, subtitle):
    ws["A1"] = "VAHDAM · Lifecycle OS"; ws["A1"].font = F_TITLE
    ws["A2"] = subtitle; ws["A2"].font = F_SUB
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30


REGIONS = ["us", "uk", "global"]
RLABEL = {"us": "US", "uk": "UK", "global": "Global"}
stats = {r: region_stats(load_catalog(r)) for r in REGIONS}

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — OVERVIEW
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = "Overview"; banner(ws, "Visual command dashboard · catalog is live data · lifecycle tabs are wired for your Supabase/Klaviyo feed")
for col, w in {"A": 22, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16}.items():
    ws.column_dimensions[col].width = w

total_products = sum(stats[r]["count"] for r in REGIONS)
all_prices = [v for r in REGIONS for v in stats[r]["prices"]]
kpis = [
    ("TOTAL ACTIVE SKUS", total_products),
    ("US CATALOG", stats["us"]["count"]),
    ("UK CATALOG", stats["uk"]["count"]),
    ("GLOBAL CATALOG", stats["global"]["count"]),
    ("AVG PRICE (ALL)", f"{round(statistics.mean(all_prices),2) if all_prices else 0}"),
    ("PRODUCT TYPES", len(set(t for r in REGIONS for t in stats[r]["types"]))),
]
ws["A4"] = "Key metrics"; ws["A4"].font = F_H
r0 = 5
for i, (label, val) in enumerate(kpis):
    col = 1 + (i % 3) * 2
    row = r0 + (i // 3) * 3
    lc = ws.cell(row=row, column=col, value=label); lc.font = F_KPI_L
    vc = ws.cell(row=row + 1, column=col, value=val); vc.font = F_KPI_N
    for rr in (row, row + 1):
        for cc in (col, col + 1):
            cell = ws.cell(row=rr, column=cc); cell.fill = FILL_KPI
            cell.border = BORDER
    ws.cell(row=row, column=col).alignment = LEFT
    ws.cell(row=row + 1, column=col).alignment = LEFT

# Products-by-region table (chart source)
tbl_r = 13
ws.cell(row=tbl_r, column=1, value="Region"); ws.cell(row=tbl_r, column=2, value="Active SKUs")
ws.cell(row=tbl_r, column=3, value="Avg Price")
style_header_row(ws, tbl_r, 3)
for i, r in enumerate(REGIONS):
    rr = tbl_r + 1 + i
    ws.cell(row=rr, column=1, value=RLABEL[r]).font = F_CELL
    ws.cell(row=rr, column=2, value=stats[r]["count"]).font = F_CELL
    ws.cell(row=rr, column=3, value=stats[r]["avg"]).font = F_CELL
    for c in range(1, 4):
        ws.cell(row=rr, column=c).border = BORDER
        if i % 2: ws.cell(row=rr, column=c).fill = FILL_ALT

bar = BarChart(); bar.type = "col"; bar.title = "Active SKUs by Region"; bar.height = 7.5; bar.width = 13
bar.style = 10
data = Reference(ws, min_col=2, min_row=tbl_r, max_row=tbl_r + len(REGIONS))
cats = Reference(ws, min_col=1, min_row=tbl_r + 1, max_row=tbl_r + len(REGIONS))
bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
bar.legend = None
ws.add_chart(bar, "A19")

# US product-type pie (top 6)
top_types = stats["us"]["types"].most_common(6)
pie_r = 13
ws.cell(row=pie_r, column=5, value="US Type"); ws.cell(row=pie_r, column=6, value="Count")
style_header_row(ws, pie_r, 0)  # no-op safe
for c in (5, 6):
    cell = ws.cell(row=pie_r, column=c); cell.fill = FILL_HEAD; cell.font = F_TH; cell.alignment = CENTER; cell.border = BORDER
for i, (t, n) in enumerate(top_types):
    rr = pie_r + 1 + i
    ws.cell(row=rr, column=5, value=t).font = F_CELL
    ws.cell(row=rr, column=6, value=n).font = F_CELL
pie = PieChart(); pie.title = "US Catalog by Product Type (Top 6)"; pie.height = 7.5; pie.width = 13
pdata = Reference(ws, min_col=6, min_row=pie_r, max_row=pie_r + len(top_types))
pcats = Reference(ws, min_col=5, min_row=pie_r + 1, max_row=pie_r + len(top_types))
pie.add_data(pdata, titles_from_data=True); pie.set_categories(pcats)
ws.add_chart(pie, "E19")

ws.cell(row=36, column=1, value="Data sources: catalog = data/catalog/*.json (built at deploy). Lifecycle/Campaign/Competitor tabs are templates — see 'Connect Live Data'.").font = F_NOTE

# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — CATALOG ANALYTICS
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Catalog Analytics"); banner(ws, "Real product catalog — counts, pricing and mix by region")
for col, w in {"A": 24, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14}.items():
    ws.column_dimensions[col].width = w
ws["A4"] = "Pricing & counts by region"; ws["A4"].font = F_H
hdr = ["Metric"] + [RLABEL[r] for r in REGIONS]
hr = 5
for c, h in enumerate(hdr, 1):
    ws.cell(row=hr, column=c, value=h)
style_header_row(ws, hr, len(hdr))
rows = [
    ("Active SKUs", "count"), ("SKUs with price", "with_price"),
    ("Min price", "min"), ("Median price", "median"),
    ("Avg price", "avg"), ("Max price", "max"),
]
for i, (label, key) in enumerate(rows):
    rr = hr + 1 + i
    ws.cell(row=rr, column=1, value=label).font = F_CELL
    for c, r in enumerate(REGIONS, 2):
        ws.cell(row=rr, column=c, value=stats[r][key]).font = F_CELL
    for c in range(1, len(hdr) + 1):
        ws.cell(row=rr, column=c).border = BORDER
        if i % 2: ws.cell(row=rr, column=c).fill = FILL_ALT

avgchart = BarChart(); avgchart.type = "col"; avgchart.title = "Average Price by Region"; avgchart.height = 7; avgchart.width = 12; avgchart.style = 12
arow = hr + 5  # 'Avg price' row
adata = Reference(ws, min_col=2, max_col=1 + len(REGIONS), min_row=arow, max_row=arow)
acats = Reference(ws, min_col=2, max_col=1 + len(REGIONS), min_row=hr, max_row=hr)
avgchart.add_data(adata, from_rows=True, titles_from_data=False); avgchart.set_categories(acats)
avgchart.legend = None
ws.add_chart(avgchart, "A14")

# Top types (US) table + bar
ws["A30"] = "US catalog — product type mix"; ws["A30"].font = F_H
tr = 31
ws.cell(row=tr, column=1, value="Product Type"); ws.cell(row=tr, column=2, value="SKUs")
style_header_row(ws, tr, 2)
for i, (t, n) in enumerate(stats["us"]["types"].most_common(10)):
    rr = tr + 1 + i
    ws.cell(row=rr, column=1, value=t).font = F_CELL
    ws.cell(row=rr, column=2, value=n).font = F_CELL
    for c in (1, 2):
        ws.cell(row=rr, column=c).border = BORDER
        if i % 2: ws.cell(row=rr, column=c).fill = FILL_ALT
tchart = BarChart(); tchart.type = "bar"; tchart.title = "US SKUs by Product Type"; tchart.height = 8; tchart.width = 12; tchart.style = 10
n_types = len(stats["us"]["types"].most_common(10))
tdata = Reference(ws, min_col=2, min_row=tr, max_row=tr + n_types)
tcats = Reference(ws, min_col=1, min_row=tr + 1, max_row=tr + n_types)
tchart.add_data(tdata, titles_from_data=True); tchart.set_categories(tcats); tchart.legend = None
ws.add_chart(tchart, "D31")

# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — PRICE DISTRIBUTION
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Price Distribution"); banner(ws, "Price-band distribution across regions (real catalog prices)")
for col, w in {"A": 18, "B": 12, "C": 12, "D": 12}.items():
    ws.column_dimensions[col].width = w
bands = [("< 15", 0, 15), ("15–25", 15, 25), ("25–35", 25, 35), ("35–50", 35, 50), ("50+", 50, 10**9)]
hr = 5
ws.cell(row=hr, column=1, value="Price Band")
for c, r in enumerate(REGIONS, 2):
    ws.cell(row=hr, column=c, value=RLABEL[r])
style_header_row(ws, hr, 1 + len(REGIONS))
for i, (label, lo, hi) in enumerate(bands):
    rr = hr + 1 + i
    ws.cell(row=rr, column=1, value=label).font = F_CELL
    for c, r in enumerate(REGIONS, 2):
        cnt = sum(1 for v in stats[r]["prices"] if lo <= v < hi)
        ws.cell(row=rr, column=c, value=cnt).font = F_CELL
    for c in range(1, 2 + len(REGIONS)):
        ws.cell(row=rr, column=c).border = BORDER
        if i % 2: ws.cell(row=rr, column=c).fill = FILL_ALT
dchart = BarChart(); dchart.type = "col"; dchart.title = "SKUs by Price Band & Region"; dchart.height = 9; dchart.width = 15; dchart.style = 11
ddata = Reference(ws, min_col=2, max_col=1 + len(REGIONS), min_row=hr, max_row=hr + len(bands))
dcats = Reference(ws, min_col=1, min_row=hr + 1, max_row=hr + len(bands))
dchart.add_data(ddata, titles_from_data=True); dchart.set_categories(dcats)
ws.add_chart(dchart, "A13")

# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — LIFECYCLE KPIs (TEMPLATE)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Lifecycle KPIs (template)"); banner(ws, "RFM cohort scorecard — TEMPLATE. Paste live rows from Supabase smart_cohorts / dashboard.")
for col, w in {"A": 20, "B": 14, "C": 14, "D": 14, "E": 14, "F": 16}.items():
    ws.column_dimensions[col].width = w
ws["A4"] = "Cohort performance (replace sample numbers with live data)"; ws["A4"].font = F_H
hdr = ["Cohort", "Customers", "Revenue", "AOV", "Orders", "Repeat Rate %"]
hr = 5
for c, h in enumerate(hdr, 1):
    ws.cell(row=hr, column=c, value=h)
style_header_row(ws, hr, len(hdr))
sample = [
    ("Champions", 0, 0, 0, 0, 0), ("Loyal", 0, 0, 0, 0, 0),
    ("Potential Loyalist", 0, 0, 0, 0, 0), ("At Risk", 0, 0, 0, 0, 0),
    ("Lapsed", 0, 0, 0, 0, 0), ("New", 0, 0, 0, 0, 0),
]
for i, row in enumerate(sample):
    rr = hr + 1 + i
    for c, val in enumerate(row, 1):
        ws.cell(row=rr, column=c, value=val).font = F_CELL
        ws.cell(row=rr, column=c).border = BORDER
        if i % 2: ws.cell(row=rr, column=c).fill = FILL_ALT
ws.cell(row=hr + 8, column=1, value="↳ Source: GET /api/brain?action=cohorts  or  Supabase table smart_cohorts. Paste values, then this chart updates automatically.").font = F_NOTE
revchart = BarChart(); revchart.type = "bar"; revchart.title = "Revenue by Cohort (live once populated)"; revchart.height = 8; revchart.width = 12; revchart.style = 10
rdata = Reference(ws, min_col=3, min_row=hr, max_row=hr + len(sample))
rcats = Reference(ws, min_col=1, min_row=hr + 1, max_row=hr + len(sample))
revchart.add_data(rdata, titles_from_data=True); revchart.set_categories(rcats); revchart.legend = None
ws.add_chart(revchart, "A15")

# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — CAMPAIGN TRACKER (TEMPLATE)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Campaign Tracker (template)"); banner(ws, "Email/SMS campaign performance — TEMPLATE. Feed from Klaviyo campaign_report or smart_generated_campaigns.")
widths = {"A": 12, "B": 12, "C": 26, "D": 12, "E": 12, "F": 10, "G": 12, "H": 10, "I": 14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
hdr = ["Date", "Channel", "Campaign", "Sent", "Opens", "Open%", "Clicks", "CTR%", "Revenue"]
hr = 5
for c, h in enumerate(hdr, 1):
    ws.cell(row=hr, column=c, value=h)
style_header_row(ws, hr, len(hdr))
# A few illustrative rows with formulas for rate columns (so visuals work on paste).
demo = [
    ("2026-06-01", "Email", "Monsoon Ritual — Winback", 12000, 3120, 480, 980),
    ("2026-06-08", "Email", "Single-Estate Spotlight", 14500, 4060, 610, 1240),
    ("2026-06-15", "SMS", "Restock — Turmeric Latte", 5200, 1560, 290, 760),
    ("2026-06-22", "Email", "Founder Note — Heritage", 13900, 4170, 700, 1510),
]
for i, (d, ch, name, sent, opens, clicks, rev) in enumerate(demo):
    rr = hr + 1 + i
    ws.cell(row=rr, column=1, value=d).font = F_CELL
    ws.cell(row=rr, column=2, value=ch).font = F_CELL
    ws.cell(row=rr, column=3, value=name).font = F_CELL
    ws.cell(row=rr, column=4, value=sent).font = F_CELL
    ws.cell(row=rr, column=5, value=opens).font = F_CELL
    ws.cell(row=rr, column=6, value=f"=IF(D{rr}=0,0,E{rr}/D{rr})").font = F_CELL
    ws.cell(row=rr, column=6).number_format = "0.0%"
    ws.cell(row=rr, column=7, value=clicks).font = F_CELL
    ws.cell(row=rr, column=8, value=f"=IF(D{rr}=0,0,G{rr}/D{rr})").font = F_CELL
    ws.cell(row=rr, column=8).number_format = "0.0%"
    ws.cell(row=rr, column=9, value=rev).font = F_CELL
    for c in range(1, len(hdr) + 1):
        ws.cell(row=rr, column=c).border = BORDER
        if i % 2: ws.cell(row=rr, column=c).fill = FILL_ALT
ws.cell(row=hr + 6, column=1, value="↳ Sample rows shown. Replace with live data; Open%/CTR are live formulas. Source: Klaviyo (op=campaign_report) or smart_generated_campaigns.").font = F_NOTE
lc = LineChart(); lc.title = "Campaign Revenue Trend"; lc.height = 8; lc.width = 14; lc.style = 12
ldata = Reference(ws, min_col=9, min_row=hr, max_row=hr + len(demo))
lcats = Reference(ws, min_col=1, min_row=hr + 1, max_row=hr + len(demo))
lc.add_data(ldata, titles_from_data=True); lc.set_categories(lcats); lc.legend = None
ws.add_chart(lc, "A13")

# ════════════════════════════════════════════════════════════════════════════
# TAB 6 — COMPETITOR (TEMPLATE)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Competitor (template)"); banner(ws, "Competitor email intel — TEMPLATE. Mirrors the competitor Google Sheet / GET /api/brain?action=benchmarks.")
widths = {"A": 18, "B": 30, "C": 30, "D": 12, "E": 12, "F": 14, "G": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
hdr = ["Brand", "Subject Line", "Offer / Angle", "Send Date", "Open%", "Channel", "Market"]
hr = 5
for c, h in enumerate(hdr, 1):
    ws.cell(row=hr, column=c, value=h)
style_header_row(ws, hr, len(hdr))
for i in range(6):
    rr = hr + 1 + i
    for c in range(1, len(hdr) + 1):
        ws.cell(row=rr, column=c, value="").border = BORDER
        if i % 2: ws.cell(row=rr, column=c).fill = FILL_ALT
ws.cell(row=hr + 8, column=1, value="↳ Source: api/_shared/competitor-core.js (Gmail IMAP → Google Sheet) or GET /api/brain?action=benchmarks.").font = F_NOTE

# ════════════════════════════════════════════════════════════════════════════
# TAB 7 — CONNECT LIVE DATA
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Connect Live Data"); banner(ws, "How to wire each template tab to live VAHDAM data")
ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 95
ws["A4"] = "Tab"; ws["B4"] = "How to feed it live data"
style_header_row(ws, 4, 2)
guide = [
    ("Lifecycle KPIs", "GET https://<deploy>/api/brain?action=cohorts (JSON) or Supabase table smart_cohorts. Paste rows; charts auto-update."),
    ("Campaign Tracker", "Klaviyo: /api/klaviyo?op=campaign_report (live once KLAVIYO_API_KEY set). Or Supabase smart_generated_campaigns."),
    ("Competitor", "api/_shared/competitor-core.js syncs Gmail IMAP → the competitor Google Sheet. Or GET /api/brain?action=benchmarks."),
    ("Catalog tabs", "Already live: rebuilt at each deploy from data/catalog/*.json (npm run build → scripts/build-catalog.js)."),
    ("Auto-refresh option", "In Google Sheets: =IMPORTDATA(\"https://<deploy>/api/...\") for CSV endpoints, or Extensions ▸ Apps Script with UrlFetchApp on a time trigger."),
    ("Re-generate this file", "python3 scripts/build_sheets_dashboard.py — rebuilds the .xlsx from the latest catalog."),
]
for i, (tab, how) in enumerate(guide):
    rr = 5 + i
    ws.cell(row=rr, column=1, value=tab).font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    ws.cell(row=rr, column=2, value=how).font = F_CELL
    for c in (1, 2):
        ws.cell(row=rr, column=c).border = BORDER; ws.cell(row=rr, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        if i % 2: ws.cell(row=rr, column=c).fill = FILL_ALT
    ws.row_dimensions[rr].height = 30

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("WROTE", OUT)
print("TABS", [s.title for s in wb.worksheets])
print("TOTALS", {RLABEL[r]: stats[r]["count"] for r in REGIONS}, "= ", total_products)
